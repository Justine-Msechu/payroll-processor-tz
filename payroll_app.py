import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys, subprocess, hashlib, json, datetime, tempfile, threading, webbrowser

# ══════════════════════════════════════════════════════════════
#  APP IDENTITY & GITHUB UPDATE SETTINGS
# ══════════════════════════════════════════════════════════════
VERSION       = "1.0.0"
APP_TITLE     = "Payroll Processor — Tanzania"
USERS_FILE    = "payroll_users.json"

# To enable auto-update:
#  1. Create a GitHub repo called  payroll-processor-tz
#  2. Upload payroll_app.py  and  version.txt  (containing just  1.0.1)
#  3. Change the two values below to your GitHub username and repo name
GITHUB_USER   = "YOUR_GITHUB_USERNAME"
GITHUB_REPO   = "payroll-processor-tz"
GITHUB_BRANCH = "main"
_base         = "https://raw.githubusercontent.com/" + GITHUB_USER + "/" + GITHUB_REPO + "/" + GITHUB_BRANCH
VERSION_URL   = _base + "/version.txt"
APP_URL       = _base + "/payroll_app.py"
UPDATE_READY  = GITHUB_USER != "YOUR_GITHUB_USERNAME"

HEADERS = [
    "Name", "Salary", "NSSF 10%", "PAY", "P.A.Y.E",
    "ALLOWANCE", "GROSS PAY", "LOAN Deduction",
    "LOAN BOARD (15%)", "NET PAY", "AMOUNT TO BE PAID",
]

# ══════════════════════════════════════════════════════════════
#  THEMES
# ══════════════════════════════════════════════════════════════
THEMES = {
    "Dark Blue": dict(
        BG="#1e2a3a", PANEL="#263547", ACCENT="#2ecc71", ACCENT2="#27ae60",
        TEXT="#ecf0f1", SUBTEXT="#95a5a6", ENTRY_BG="#2d4059", ENTRY_FG="#ecf0f1",
        RED="#e74c3c", GOLD="#f39c12", HEADER_BG="#1a5276",
        BTN_SAVE="#2980b9", BTN_PRINT="#e67e22", BTN_OPEN="#8e44ad",
        TREE_BG="#2d4059", TREE_FG="#ecf0f1", TREE_HEAD="#1A5276", SEP="#34495e",
        MENU_BG="#1e2a3a", MENU_FG="#ecf0f1", MENU_ACT="#27ae60",
    ),
    "Light": dict(
        BG="#f0f4f8", PANEL="#ffffff", ACCENT="#2563eb", ACCENT2="#1d4ed8",
        TEXT="#1e293b", SUBTEXT="#64748b", ENTRY_BG="#e2e8f0", ENTRY_FG="#1e293b",
        RED="#dc2626", GOLD="#d97706", HEADER_BG="#2563eb",
        BTN_SAVE="#0369a1", BTN_PRINT="#b45309", BTN_OPEN="#7c3aed",
        TREE_BG="#f8fafc", TREE_FG="#1e293b", TREE_HEAD="#1e40af", SEP="#cbd5e1",
        MENU_BG="#ffffff", MENU_FG="#1e293b", MENU_ACT="#2563eb",
    ),
    "Green": dict(
        BG="#0f2418", PANEL="#1a3a2a", ACCENT="#4ade80", ACCENT2="#16a34a",
        TEXT="#dcfce7", SUBTEXT="#86efac", ENTRY_BG="#14532d", ENTRY_FG="#dcfce7",
        RED="#f87171", GOLD="#fbbf24", HEADER_BG="#14532d",
        BTN_SAVE="#0e7490", BTN_PRINT="#b45309", BTN_OPEN="#7c3aed",
        TREE_BG="#14532d", TREE_FG="#dcfce7", TREE_HEAD="#166534", SEP="#166534",
        MENU_BG="#0f2418", MENU_FG="#dcfce7", MENU_ACT="#16a34a",
    ),
    "Purple": dict(
        BG="#1e1b2e", PANEL="#2d2b45", ACCENT="#a78bfa", ACCENT2="#7c3aed",
        TEXT="#ede9fe", SUBTEXT="#c4b5fd", ENTRY_BG="#3b3563", ENTRY_FG="#ede9fe",
        RED="#f87171", GOLD="#fbbf24", HEADER_BG="#3b3563",
        BTN_SAVE="#2563eb", BTN_PRINT="#d97706", BTN_OPEN="#16a34a",
        TREE_BG="#3b3563", TREE_FG="#ede9fe", TREE_HEAD="#4c1d95", SEP="#4c1d95",
        MENU_BG="#1e1b2e", MENU_FG="#ede9fe", MENU_ACT="#7c3aed",
    ),
    "Orange": dict(
        BG="#1c1007", PANEL="#2d1f0e", ACCENT="#fb923c", ACCENT2="#ea580c",
        TEXT="#fff7ed", SUBTEXT="#fed7aa", ENTRY_BG="#431407", ENTRY_FG="#fff7ed",
        RED="#f87171", GOLD="#fbbf24", HEADER_BG="#431407",
        BTN_SAVE="#0369a1", BTN_PRINT="#7c3aed", BTN_OPEN="#16a34a",
        TREE_BG="#431407", TREE_FG="#fff7ed", TREE_HEAD="#7c2d12", SEP="#7c2d12",
        MENU_BG="#1c1007", MENU_FG="#fff7ed", MENU_ACT="#ea580c",
    ),
}

# ══════════════════════════════════════════════════════════════
#  CALCULATIONS
# ══════════════════════════════════════════════════════════════
def calc_paye_tz(pay):
    if pay <= 270_000:   return 0.0
    if pay <= 520_000:   return round((pay - 270_000) * 0.08, 2)
    if pay <= 760_000:   return round(20_000 + (pay - 520_000) * 0.20, 2)
    if pay <= 1_000_000: return round(68_000 + (pay - 760_000) * 0.25, 2)
    return round(128_000 + (pay - 1_000_000) * 0.30, 2)

def calculate(salary, allowance, has_loan, loan_amount, has_loan_board):
    nssf        = round(salary * 0.10, 2)
    pay         = round(salary - nssf, 2)
    paye        = calc_paye_tz(pay)
    gross_pay   = round(salary + allowance, 2)
    loan_ded    = round(loan_amount, 2) if has_loan else 0.0
    loan_board  = round(salary * 0.15, 2) if has_loan_board else 0.0
    net_pay     = round(gross_pay - nssf - paye, 2)
    amount_paid = round(net_pay - loan_ded - loan_board, 2)
    return nssf, pay, paye, gross_pay, loan_ded, loan_board, net_pay, amount_paid

# ══════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════
def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

# Developer/admin secret credentials — never shown to end users
DEVELOPER_USER = "devadmin"
DEVELOPER_PASS = "Tz@Payroll!Dev99"   # change this to your own secret password

def load_users():
    """Load user accounts. The developer account is kept separately."""
    if not os.path.exists(USERS_FILE):
        save_users({})
        return {}
    with open(USERS_FILE) as f:
        return json.load(f)

def has_any_users():
    """Return True if at least one regular user account exists."""
    return len(load_users()) > 0

def verify_login(username, password):
    """
    Check credentials. Developer account checked first (silent).
    Regular users checked from file.
    """
    u = username.strip().lower()
    p = password
    # Developer account — hidden, never shown in UI
    if u == DEVELOPER_USER and hash_pw(p) == hash_pw(DEVELOPER_PASS):
        return "admin"
    # Regular user accounts
    users = load_users()
    rec = users.get(u)
    if rec and rec["hash"] == hash_pw(p):
        return rec["role"]
    return None

def save_users(u):
    with open(USERS_FILE, "w") as f:
        json.dump(u, f, indent=2)



# ══════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════
def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def save_to_excel(records, filepath, created_by, month_label):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title  = "Payroll"
    now_str   = datetime.datetime.now().strftime("%d %B %Y  %H:%M")
    info_fill = PatternFill("solid", fgColor="D6EAF8")

    # Title row
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value     = "PAYROLL REGISTER  —  " + month_label
    c.font      = Font(bold=True, size=13, color="1A5276")
    c.fill      = info_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for row_i, (label, value) in enumerate([
        ("Created by",          created_by),
        ("Date & Time",         now_str),
        ("Number of Employees", str(len(records))),
    ], start=2):
        ws.cell(row=row_i, column=1, value=label).font = Font(bold=True, size=9, color="1A5276")
        ws.cell(row=row_i, column=2, value=value).font = Font(size=9, color="2C3E50")
        ws.cell(row=row_i, column=1).fill = info_fill
        ws.cell(row=row_i, column=2).fill = info_fill

    ws.row_dimensions[5].height = 6

    # Column headers
    header_row = 6
    hfill = PatternFill("solid", fgColor="1A5276")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[header_row].height = 32

    # Data rows
    next_row = header_row + 1
    for r in records:
        row_data = [r["name"], r["salary"], r["nssf"], r["pay"], r["paye"],
                    r["allowance"], r["gross_pay"], r["loan_ded"],
                    r["loan_board"], r["net_pay"], r["amount_paid"]]
        rfill = PatternFill("solid", fgColor="EBF5FB" if next_row % 2 == 0 else "FDFEFE")
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=next_row, column=col, value=val)
            c.fill      = rfill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border()
            if col > 1:
                c.number_format = "#,##0.00"
        next_row += 1

    ws.row_dimensions[next_row].height = 10
    next_row += 1

    # Summary box
    total_salary = sum(r["salary"]      for r in records)
    total_nssf   = sum(r["nssf"]        for r in records)
    total_paye   = sum(r["paye"]        for r in records)
    total_gross  = sum(r["gross_pay"]   for r in records)
    total_loan   = sum(r["loan_ded"]    for r in records)
    total_lb     = sum(r["loan_board"]  for r in records)
    total_net    = sum(r["net_pay"]     for r in records)
    total_amount = sum(r["amount_paid"] for r in records)

    sum_fill = PatternFill("solid", fgColor="1A5276")
    key_fill = PatternFill("solid", fgColor="D6EAF8")
    hl_fill  = PatternFill("solid", fgColor="145A32")

    ws.merge_cells("A" + str(next_row) + ":K" + str(next_row))
    sc = ws.cell(row=next_row, column=1, value="PAYROLL SUMMARY")
    sc.font      = Font(bold=True, color="FFFFFF", size=11)
    sc.fill      = sum_fill
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[next_row].height = 22
    next_row += 1

    for label, value, hl in [
        ("Total Number of Employees",       len(records),  False),
        ("Total Gross Salary",               total_salary,  False),
        ("Total NSSF Deductions",            total_nssf,    False),
        ("Total PAYE Tax",                   total_paye,    False),
        ("Total Gross Pay (Salary+Allow.)",  total_gross,   False),
        ("Total Loan Deductions",            total_loan,    False),
        ("Total Loan Board Deductions",      total_lb,      False),
        ("Total Net Pay",                    total_net,     False),
        ("TOTAL AMOUNT TO BE PAID OUT",      total_amount,  True),
    ]:
        lc = ws.cell(row=next_row, column=1, value=label)
        vc = ws.cell(row=next_row, column=2,
                     value=value if isinstance(value, int) else round(value, 2))
        fill = hl_fill if hl else key_fill
        for cell in (lc, vc):
            cell.fill   = fill
            cell.border = thin_border()
        if hl:
            lc.font = Font(bold=True, color="FFFFFF", size=11)
            vc.font = Font(bold=True, color="FFFFFF", size=11)
            ws.row_dimensions[next_row].height = 24
        else:
            lc.font = Font(bold=True, size=9, color="1A5276")
            vc.font = Font(size=9, color="1A5276")
        lc.alignment = Alignment(horizontal="left",  vertical="center", indent=1)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        if not isinstance(value, int):
            vc.number_format = "#,##0.00"
        next_row += 1

    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(w + 4, 18)

    ws.freeze_panes = "A" + str(header_row + 1)
    wb.save(filepath)

# ══════════════════════════════════════════════════════════════
#  PRINT — HTML in browser
# ══════════════════════════════════════════════════════════════
def make_print_html(records, month_label, created_by):
    now_str = datetime.datetime.now().strftime("%d %B %Y  %H:%M")
    rows_html = ""
    for i, r in enumerate(records):
        bg = "#eaf4fb" if i % 2 == 0 else "#ffffff"
        rows_html += (
            "<tr style='background:" + bg + "'>"
            "<td style='text-align:left;font-weight:500'>" + r["name"] + "</td>"
            "<td>" + "{:,.2f}".format(r["salary"])      + "</td>"
            "<td>" + "{:,.2f}".format(r["nssf"])        + "</td>"
            "<td>" + "{:,.2f}".format(r["pay"])         + "</td>"
            "<td>" + "{:,.2f}".format(r["paye"])        + "</td>"
            "<td>" + "{:,.2f}".format(r["allowance"])   + "</td>"
            "<td>" + "{:,.2f}".format(r["gross_pay"])   + "</td>"
            "<td>" + "{:,.2f}".format(r["loan_ded"])    + "</td>"
            "<td>" + "{:,.2f}".format(r["loan_board"])  + "</td>"
            "<td>" + "{:,.2f}".format(r["net_pay"])     + "</td>"
            "<td style='color:#145a32;font-weight:bold'>" + "{:,.2f}".format(r["amount_paid"]) + "</td>"
            "</tr>"
        )

    ts = sum(r["salary"]      for r in records)
    tn = sum(r["nssf"]        for r in records)
    tp = sum(r["paye"]        for r in records)
    tt = sum(r["net_pay"]     for r in records)
    ta = sum(r["amount_paid"] for r in records)

    header_cells = "".join("<th>" + h + "</th>" for h in HEADERS)

    return (
        "<!DOCTYPE html><html><head><meta charset='UTF-8'>"
        "<title>Payroll " + month_label + "</title>"
        "<style>"
        "@page{size:A4 landscape;margin:1.2cm}"
        "*{box-sizing:border-box;font-family:'Segoe UI',Arial,sans-serif;margin:0;padding:0}"
        "body{padding:14px;color:#1e293b}"
        ".top{border-bottom:3px solid #1A5276;padding-bottom:8px;margin-bottom:10px;"
        "display:flex;justify-content:space-between;align-items:flex-end}"
        ".top h1{font-size:18px;color:#1A5276}"
        ".meta{font-size:10px;color:#64748b;text-align:right;line-height:1.7}"
        "table{width:100%;border-collapse:collapse;font-size:9px}"
        "th{background:#1A5276;color:white;padding:7px 4px;text-align:center;"
        "font-weight:bold;border:1px solid #b0c4d8}"
        "td{padding:5px 4px;text-align:right;border:1px solid #dde3ea;white-space:nowrap}"
        ".summary{margin-top:16px;display:flex;gap:10px;flex-wrap:wrap}"
        ".sc{background:#f0f9ff;border:2px solid #1A5276;border-radius:6px;"
        "padding:10px 16px;min-width:150px;text-align:center}"
        ".sc.hl{background:#145a32;border-color:#145a32}"
        ".sc .lb{font-size:9px;color:#64748b;margin-bottom:4px}"
        ".sc.hl .lb{color:#a7f3d0}"
        ".sc .vl{font-size:13px;font-weight:bold;color:#1A5276}"
        ".sc.hl .vl{color:white;font-size:15px}"
        ".sigs{margin-top:20px;display:flex;justify-content:space-around}"
        ".sig{text-align:center;width:180px}"
        ".sig-line{border-top:1px solid #555;margin-bottom:3px}"
        ".footer{margin-top:10px;font-size:8px;color:#94a3b8;text-align:center}"
        "@media print{body{padding:0}}"
        "</style></head><body>"
        "<div class='top'>"
        "<div><h1>PAYROLL REGISTER" + ("  \u2014  " + month_label if month_label else "") + "</h1></div>"
        "<div class='meta'>"
        "<div><b>Created by:</b> " + created_by + "</div>"
        "<div><b>Date &amp; Time:</b> " + now_str + "</div>"
        "<div><b>Employees:</b> " + str(len(records)) + "</div>"
        "</div></div>"
        "<table><thead><tr>" + header_cells + "</tr></thead>"
        "<tbody>" + rows_html + "</tbody></table>"
        "<div class='summary'>"
        "<div class='sc'><div class='lb'>Employees</div><div class='vl'>" + str(len(records)) + "</div></div>"
        "<div class='sc'><div class='lb'>Total Salary</div><div class='vl'>TZS " + "{:,.2f}".format(ts) + "</div></div>"
        "<div class='sc'><div class='lb'>Total NSSF</div><div class='vl'>TZS " + "{:,.2f}".format(tn) + "</div></div>"
        "<div class='sc'><div class='lb'>Total PAYE</div><div class='vl'>TZS " + "{:,.2f}".format(tp) + "</div></div>"
        "<div class='sc'><div class='lb'>Total Net Pay</div><div class='vl'>TZS " + "{:,.2f}".format(tt) + "</div></div>"
        "<div class='sc hl'><div class='lb'>TOTAL AMOUNT TO BE PAID OUT</div>"
        "<div class='vl'>TZS " + "{:,.2f}".format(ta) + "</div></div>"
        "</div>"
        "<div class='sigs'>"
        "<div class='sig'><div class='sig-line'></div>Prepared by</div>"
        "<div class='sig'><div class='sig-line'></div>Reviewed by</div>"
        "<div class='sig'><div class='sig-line'></div>Approved by</div>"
        "</div>"
        "<div class='footer'>All amounts in TZS &nbsp;|&nbsp; "
        "NSSF = 10% of Salary &nbsp;|&nbsp; PAYE = TRA 2024/25 &nbsp;|&nbsp; "
        "Loan Board = 15% of Salary</div>"
        "</body></html>"
    )

def open_print_in_browser(html_content):
    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix=".html", mode="w", encoding="utf-8")
    tmp.write(html_content)
    tmp.close()
    webbrowser.open("file:///" + tmp.name.replace("\\", "/"))

# ══════════════════════════════════════════════════════════════
#  AUTO-UPDATE  (compares actual code content, not version number)
# ══════════════════════════════════════════════════════════════
def _file_hash(text):
    """MD5 of file content — used to detect any code change on GitHub."""
    import hashlib as _hl
    return _hl.md5(text.encode("utf-8")).hexdigest()

def fetch_update_in_bg(holder):
    """
    Background thread.
    holder[0] will be set to one of:
      None          — still running
      "same"        — no changes found
      "error"       — network / other error
      <str content> — the new source code (different from local)
    """
    def _run():
        try:
            import urllib.request
            with urllib.request.urlopen(APP_URL, timeout=10) as r:
                remote_src = r.read().decode("utf-8")
            this_file = os.path.abspath(__file__)
            with open(this_file, "r", encoding="utf-8") as f:
                local_src = f.read()
            if _file_hash(remote_src) != _file_hash(local_src):
                holder[0] = remote_src   # changed — return new code
            else:
                holder[0] = "same"
        except Exception:
            holder[0] = "error"
    threading.Thread(target=_run, daemon=True).start()

def apply_update(new_source, parent):
    """Write new_source to disk and restart the app."""
    import shutil
    try:
        this_file = os.path.abspath(__file__)
        backup    = this_file + ".bak"
        shutil.copy2(this_file, backup)
        with open(this_file, "w", encoding="utf-8") as f:
            f.write(new_source)
        messagebox.showinfo("Updated!",
            "The app has been updated successfully.\nIt will now restart.",
            parent=parent)
        os.execv(sys.executable, [sys.executable, this_file])
    except Exception as e:
        messagebox.showerror("Update Failed",
            "Could not apply the update.\n\nError: " + str(e),
            parent=parent)


# ══════════════════════════════════════════════════════════════
#  SCROLLABLE FRAME
# ══════════════════════════════════════════════════════════════
class ScrollableFrame(tk.Frame):
    def __init__(self, parent, bg):
        super().__init__(parent, bg=bg)
        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(self.canvas, bg=bg)
        self._win  = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", lambda e: self.canvas.configure(
            scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(
            self._win, width=e.width))
        for w in (self.canvas, self.inner):
            w.bind("<MouseWheel>", lambda e: self.canvas.yview_scroll(
                int(-1 * (e.delta / 120)), "units"))

    def set_bg(self, bg):
        self.canvas.configure(bg=bg)
        self.inner.configure(bg=bg)

# ══════════════════════════════════════════════════════════════
#  TOOLTIP
# ══════════════════════════════════════════════════════════════
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text   = text
        self.tip    = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, _=None):
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self.tip = tk.Toplevel(self.widget)
        self.tip.wm_overrideredirect(True)
        self.tip.wm_geometry("+" + str(x) + "+" + str(y))
        tk.Label(self.tip, text=self.text, font=("Segoe UI", 8),
                 bg="#fffde7", fg="#333", relief="solid", bd=1,
                 padx=6, pady=3).pack()

    def hide(self, _=None):
        if self.tip:
            self.tip.destroy()
            self.tip = None

# ══════════════════════════════════════════════════════════════
#  FIRST SETUP WINDOW — shown only on very first launch
# ══════════════════════════════════════════════════════════════
class FirstSetupWindow(tk.Tk):
    """
    Shown the very first time the app is opened (no user accounts yet).
    Asks the person to create their own account.
    Simple — just name, username, password, confirm password.
    """
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE + "  —  Welcome!")
        self.resizable(False, False)
        self.configure(bg="#1e2a3a")
        self.created_user = None
        self.created_role = None
        self._build()
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(str(w)+"x"+str(h)+"+"+str((sw-w)//2)+"+"+str((sh-h)//2))

    def _build(self):
        T = THEMES["Dark Blue"]

        # Header
        top = tk.Frame(self, bg="#1a5276", pady=20)
        top.pack(fill="x")
        tk.Label(top, text="💼", font=("Segoe UI",40),
                 bg="#1a5276", fg="white").pack()
        tk.Label(top, text="Welcome to Payroll Processor",
                 font=("Segoe UI",16,"bold"), bg="#1a5276", fg="white").pack()
        tk.Label(top, text="Tanzania  ·  Let's set up your account to get started",
                 font=("Segoe UI",9), bg="#1a5276", fg="#d5f5e3").pack(pady=(2,0))

        # Card
        card = tk.Frame(self, bg=T["PANEL"], padx=44, pady=28)
        card.pack(padx=44, pady=24)

        tk.Label(card, text="Create Your Account",
                 font=("Segoe UI",12,"bold"), bg=T["PANEL"], fg=T["ACCENT"]
                 ).grid(row=0, column=0, sticky="w", pady=(0,14))

        self._name_var  = tk.StringVar()
        self._user_var  = tk.StringVar()
        self._pass_var  = tk.StringVar()
        self._conf_var  = tk.StringVar()

        fields = [
            ("👤  Your Full Name",  self._name_var, "", "e.g.  Justine Msechu"),
            ("🔤  Choose Username",  self._user_var, "", "e.g.  justine  (no spaces)"),
            ("🔑  Choose Password",  self._pass_var, "●", "At least 6 characters"),
            ("🔑  Confirm Password", self._conf_var, "●", "Type the same password again"),
        ]
        for row_i, (label, var, show, hint) in enumerate(fields, start=1):
            tk.Label(card, text=label, font=("Segoe UI",10,"bold"),
                     bg=T["PANEL"], fg=T["TEXT"], anchor="w"
                     ).grid(row=row_i*3-2, column=0, sticky="w", pady=(8,1))
            e = tk.Entry(card, textvariable=var, show=show, font=("Segoe UI",11),
                         bg=T["ENTRY_BG"], fg=T["ENTRY_FG"],
                         insertbackground="white", relief="flat", bd=5, width=30)
            e.grid(row=row_i*3-1, column=0, ipady=6, sticky="ew")
            tk.Label(card, text="   " + hint, font=("Segoe UI",7),
                     bg=T["PANEL"], fg=T["SUBTEXT"]
                     ).grid(row=row_i*3, column=0, sticky="w")
            if row_i == 1:
                e.focus_set()

        self.err = tk.StringVar()
        tk.Label(card, textvariable=self.err, font=("Segoe UI",9),
                 bg=T["PANEL"], fg=T["RED"], wraplength=320
                 ).grid(row=13, column=0, pady=(8,0))

        tk.Button(card, text="✅   Create Account and Start",
                  font=("Segoe UI",12,"bold"),
                  bg=T["ACCENT2"], fg="white", relief="flat", cursor="hand2",
                  command=self._create, activebackground=T["ACCENT"]
                  ).grid(row=14, column=0, pady=(16,0), ipadx=12, ipady=10, sticky="ew")

        tk.Label(card,
                 text="You can add more users later from inside the app.",
                 font=("Segoe UI",7,"italic"), bg=T["PANEL"], fg=T["SUBTEXT"]
                 ).grid(row=15, column=0, pady=(10,0))

        self.bind("<Return>", lambda e: self._create())

    def _create(self):
        name  = self._name_var.get().strip()
        uname = self._user_var.get().strip().lower().replace(" ", "")
        pw    = self._pass_var.get()
        conf  = self._conf_var.get()

        if not name:
            self.err.set("⚠  Please enter your full name.")
            return
        if not uname:
            self.err.set("⚠  Please choose a username.")
            return
        if len(uname) < 3:
            self.err.set("⚠  Username must be at least 3 characters.")
            return
        if uname == DEVELOPER_USER:
            self.err.set("⚠  That username is reserved. Please choose another.")
            return
        if len(pw) < 6:
            self.err.set("⚠  Password must be at least 6 characters.")
            return
        if pw != conf:
            self.err.set("⚠  Passwords do not match. Please try again.")
            return

        # Save the new account as accountant
        users = load_users()
        users[uname] = {
            "hash":      hash_pw(pw),
            "role":      "accountant",
            "full_name": name,
        }
        save_users(users)
        self.created_user = uname
        self.created_role = "accountant"
        self.destroy()


# ══════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ══════════════════════════════════════════════════════════════
class LoginWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.resizable(False, False)
        self.configure(bg="#1e2a3a")
        self.logged_in_user = None
        self.logged_in_role = None
        self._build()
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(str(w) + "x" + str(h) + "+" + str((sw-w)//2) + "+" + str((sh-h)//2))

    def _build(self):
        T = THEMES["Dark Blue"]
        top = tk.Frame(self, bg="#1a5276", pady=22)
        top.pack(fill="x")
        tk.Label(top, text="💼",  font=("Segoe UI",40), bg="#1a5276", fg="white").pack()
        tk.Label(top, text="PAYROLL PROCESSOR",
                 font=("Segoe UI",17,"bold"), bg="#1a5276", fg="white").pack()
        tk.Label(top, text="Tanzania  ·  Please log in to continue",
                 font=("Segoe UI",9), bg="#1a5276", fg="#d5f5e3").pack(pady=(2,0))

        card = tk.Frame(self, bg=T["PANEL"], padx=44, pady=30)
        card.pack(padx=50, pady=28)
        self._u = tk.StringVar()
        self._p = tk.StringVar()
        for row, (icon, lbl, var, show) in enumerate([
            ("👤", "Username", self._u, ""),
            ("🔑", "Password", self._p, "●"),
        ], start=1):
            tk.Label(card, text=icon + "  " + lbl, font=("Segoe UI",10),
                     bg=T["PANEL"], fg=T["TEXT"], anchor="w"
                     ).grid(row=row*2-1, column=0, sticky="w", pady=(8,1))
            e = tk.Entry(card, textvariable=var, show=show, font=("Segoe UI",12),
                         bg=T["ENTRY_BG"], fg=T["ENTRY_FG"],
                         insertbackground="white", relief="flat", bd=6, width=24)
            e.grid(row=row*2, column=0, ipady=6, sticky="ew")
            if row == 1: e.focus_set()
        self.err = tk.StringVar()
        tk.Label(card, textvariable=self.err, font=("Segoe UI",9),
                 bg=T["PANEL"], fg=T["RED"], wraplength=280
                 ).grid(row=5, column=0, pady=(6,0))
        tk.Button(card, text="🔓   LOG IN", font=("Segoe UI",12,"bold"),
                  bg=T["ACCENT2"], fg="white", relief="flat", cursor="hand2",
                  command=self._login, activebackground=T["ACCENT"]
                  ).grid(row=6, column=0, pady=(16,0), ipadx=12, ipady=10, sticky="ew")
        # No hint shown — admin is a hidden developer account
        self.bind("<Return>", lambda e: self._login())

    def _login(self):
        u, p = self._u.get().strip(), self._p.get()
        if not u or not p:
            self.err.set("⚠  Please fill in both fields.")
            return
        role = verify_login(u, p)
        if role:
            self.logged_in_user = u.lower()
            self.logged_in_role = role
            self.destroy()
        else:
            self.err.set("❌  Wrong username or password.")
            self._p.set("")

# ══════════════════════════════════════════════════════════════
#  USER MANAGER
# ══════════════════════════════════════════════════════════════
class UserManagerDialog(tk.Toplevel):
    def __init__(self, parent, T):
        super().__init__(parent)
        self.T = T
        self.title("Manage Users")
        self.configure(bg=T["BG"])
        self.resizable(False, False)
        self._build()
        self.grab_set()
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        px = parent.winfo_x() + parent.winfo_width()//2
        py = parent.winfo_y() + parent.winfo_height()//2
        self.geometry(str(w)+"x"+str(h)+"+"+str(px-w//2)+"+"+str(py-h//2))

    def _build(self):
        T = self.T
        tk.Label(self, text="👥  Manage Users", font=("Segoe UI",13,"bold"),
                 bg=T["BG"], fg=T["ACCENT"]).pack(pady=(16,4))
        tk.Frame(self, bg=T["ACCENT"], height=2).pack(fill="x", padx=20)
        lf = tk.Frame(self, bg=T["PANEL"], padx=14, pady=10)
        lf.pack(fill="both", expand=True, padx=20, pady=10)
        tk.Label(lf, text="Existing Users:", font=("Segoe UI",9,"bold"),
                 bg=T["PANEL"], fg=T["TEXT"]).pack(anchor="w")
        self.lb = tk.Listbox(lf, font=("Segoe UI",9), bg=T["ENTRY_BG"],
                              fg=T["ENTRY_FG"], relief="flat", height=5,
                              selectbackground=T["ACCENT2"])
        self.lb.pack(fill="both", expand=True, pady=4)
        tk.Button(lf, text="🗑  Remove Selected", font=("Segoe UI",8),
                  bg=T["RED"], fg="white", relief="flat", cursor="hand2",
                  command=self._remove).pack(anchor="e", ipady=3, ipadx=8, pady=(2,0))
        self._refresh()
        af = tk.Frame(self, bg=T["PANEL"], padx=14, pady=10)
        af.pack(fill="x", padx=20, pady=(0,10))
        tk.Label(af, text="Add New User:", font=("Segoe UI",9,"bold"),
                 bg=T["PANEL"], fg=T["TEXT"]).pack(anchor="w", pady=(0,6))
        self._nu = tk.StringVar()
        self._np = tk.StringVar()
        self._nr = tk.StringVar(value="accountant")
        for lbl, var, show in [("Username",self._nu,""),("Password",self._np,"●")]:
            row = tk.Frame(af, bg=T["PANEL"]); row.pack(fill="x", pady=2)
            tk.Label(row, text=lbl, width=12, anchor="w", font=("Segoe UI",9),
                     bg=T["PANEL"], fg=T["TEXT"]).pack(side="left")
            tk.Entry(row, textvariable=var, show=show, font=("Segoe UI",9),
                     bg=T["ENTRY_BG"], fg=T["ENTRY_FG"], insertbackground="white",
                     relief="flat", bd=3).pack(side="right", expand=True, fill="x", ipady=4)
        rr = tk.Frame(af, bg=T["PANEL"]); rr.pack(fill="x", pady=2)
        tk.Label(rr, text="Role", width=12, anchor="w", font=("Segoe UI",9),
                 bg=T["PANEL"], fg=T["TEXT"]).pack(side="left")
        om = tk.OptionMenu(rr, self._nr, "accountant")
        om.configure(bg=T["ENTRY_BG"], fg=T["TEXT"], relief="flat",
                     font=("Segoe UI",9), highlightthickness=0)
        om["menu"].configure(bg=T["ENTRY_BG"], fg=T["TEXT"])
        om.pack(side="left", padx=2)
        tk.Button(af, text="➕  Add User", font=("Segoe UI",10,"bold"),
                  bg=T["ACCENT2"], fg="white", relief="flat", cursor="hand2",
                  command=self._add).pack(fill="x", pady=(8,0), ipady=6)
        self.msg = tk.StringVar()
        tk.Label(self, textvariable=self.msg, font=("Segoe UI",8),
                 bg=T["BG"], fg=T["GOLD"]).pack(pady=(0,10))

    def _refresh(self):
        self.lb.delete(0, "end")
        for u, d in load_users().items():
            # Never show the developer account in the user list
            if u == DEVELOPER_USER:
                continue
            self.lb.insert("end", "  " + u + "   (" + d["role"] + ")")

    def _add(self):
        u, p = self._nu.get().strip().lower(), self._np.get()
        if not u or not p:
            self.msg.set("⚠  Enter both username and password.")
            return
        users = load_users()
        if u in users:
            self.msg.set("⚠  User '" + u + "' already exists.")
            return
        users[u] = {"hash": hash_pw(p), "role": self._nr.get()}
        save_users(users)
        self.msg.set("✅  User '" + u + "' added.")
        self._nu.set(""); self._np.set(""); self._refresh()

    def _remove(self):
        sel = self.lb.curselection()
        if not sel: return
        uname = self.lb.get(sel[0]).strip().split()[0]
        if uname == DEVELOPER_USER:
            messagebox.showwarning("Cannot Remove",
                "This is a system account and cannot be removed.", parent=self)
            return
        if len([u for u in load_users() if u != DEVELOPER_USER]) == 1:
            messagebox.showwarning("Cannot Remove",
                "Cannot remove the only user account.", parent=self)
            return
        if messagebox.askyesno("Confirm", "Remove user '" + uname + "'?", parent=self):
            users = load_users(); users.pop(uname, None); save_users(users)
            self.msg.set("User '" + uname + "' removed."); self._refresh()

# ══════════════════════════════════════════════════════════════
#  CHANGE PASSWORD
# ══════════════════════════════════════════════════════════════
class ChangePasswordDialog(tk.Toplevel):
    def __init__(self, parent, T, current_user):
        super().__init__(parent)
        self.T = T; self.current_user = current_user
        self.title("Change My Password")
        self.configure(bg=T["BG"]); self.resizable(False, False)
        self._build(); self.grab_set()
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        px = parent.winfo_x() + parent.winfo_width()//2
        py = parent.winfo_y() + parent.winfo_height()//2
        self.geometry(str(w)+"x"+str(h)+"+"+str(px-w//2)+"+"+str(py-h//2))

    def _build(self):
        T = self.T
        card = tk.Frame(self, bg=T["PANEL"], padx=32, pady=26)
        card.pack(padx=26, pady=26)
        tk.Label(card, text="🔒  Change My Password",
                 font=("Segoe UI",12,"bold"), bg=T["PANEL"], fg=T["ACCENT"]
                 ).grid(row=0, column=0, columnspan=2, pady=(0,16))
        self._vars = {}
        for r, (lbl, key) in enumerate([
            ("Current Password","old"), ("New Password","new1"), ("Confirm New","new2")
        ], start=1):
            tk.Label(card, text=lbl, font=("Segoe UI",9), bg=T["PANEL"],
                     fg=T["TEXT"], anchor="w", width=18).grid(row=r, column=0, sticky="w", pady=5)
            var = tk.StringVar()
            tk.Entry(card, textvariable=var, show="●", font=("Segoe UI",10),
                     bg=T["ENTRY_BG"], fg=T["ENTRY_FG"],
                     insertbackground="white", relief="flat", bd=4, width=22
                     ).grid(row=r, column=1, padx=(8,0), ipady=6)
            self._vars[key] = var
        self.msg = tk.StringVar()
        tk.Label(card, textvariable=self.msg, font=("Segoe UI",8),
                 bg=T["PANEL"], fg=T["RED"], wraplength=260
                 ).grid(row=4, column=0, columnspan=2, pady=(4,0))
        tk.Button(card, text="✅  Save New Password",
                  font=("Segoe UI",10,"bold"), bg=T["ACCENT2"], fg="white",
                  relief="flat", cursor="hand2", command=self._change
                  ).grid(row=5, column=0, columnspan=2, pady=(14,0),
                         ipadx=10, ipady=8, sticky="ew")

    def _change(self):
        old, n1, n2 = self._vars["old"].get(), self._vars["new1"].get(), self._vars["new2"].get()
        users = load_users(); u = users.get(self.current_user)
        if not u or u["hash"] != hash_pw(old):
            self.msg.set("⚠  Current password is incorrect."); return
        if len(n1) < 6:
            self.msg.set("⚠  New password must be at least 6 characters."); return
        if n1 != n2:
            self.msg.set("⚠  Passwords do not match."); return
        users[self.current_user]["hash"] = hash_pw(n1); save_users(users)
        messagebox.showinfo("Done ✅", "Your password has been updated!", parent=self)
        self.destroy()

# ══════════════════════════════════════════════════════════════
#  EDIT EMPLOYEE DIALOG
# ══════════════════════════════════════════════════════════════
class EditEmployeeDialog(tk.Toplevel):
    """
    Opens when user double-clicks a row.
    Pre-fills all fields with current values.
    On Save, recalculates everything and updates the record in place.
    """
    def __init__(self, parent, T, record, on_save):
        super().__init__(parent)
        self.T       = T
        self.record  = record        # dict — the existing employee record
        self.on_save = on_save       # callback(updated_record)
        self.title("✏  Edit Employee")
        self.configure(bg=T["BG"])
        self.resizable(False, False)
        self._build()
        self.grab_set()
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        px = parent.winfo_x() + parent.winfo_width()  // 2
        py = parent.winfo_y() + parent.winfo_height() // 2
        self.geometry(str(w)+"x"+str(h)+"+"+str(px-w//2)+"+"+str(py-h//2))

    def _build(self):
        T  = self.T
        r  = self.record

        # Header
        hdr = tk.Frame(self, bg=T["HEADER_BG"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="✏  Edit Employee Details",
                 font=("Segoe UI",12,"bold"), bg=T["HEADER_BG"], fg="white").pack()
        tk.Label(hdr, text="Change what you need, then click  Save Changes",
                 font=("Segoe UI",8), bg=T["HEADER_BG"], fg="white").pack()

        card = tk.Frame(self, bg=T["PANEL"], padx=24, pady=18)
        card.pack(fill="both", expand=True, padx=16, pady=12)

        self._name_var      = tk.StringVar(value=r["name"])
        self._salary_var    = tk.StringVar(value=str(r["salary"]))
        self._allow_var     = tk.StringVar(value=str(r["allowance"]))
        self._loan_var      = tk.StringVar(value=str(r["loan_ded"]) if r["loan_ded"] > 0 else "")
        self._has_loan      = tk.BooleanVar(value=r["loan_ded"] > 0)
        self._has_lb        = tk.BooleanVar(value=r["loan_board"] > 0)

        # Input rows
        for row_i, (label, var, show) in enumerate([
            ("👤  Employee Name",      self._name_var,   ""),
            ("💰  Basic Salary (TZS)", self._salary_var, ""),
            ("➕  Allowance (TZS)",    self._allow_var,  ""),
        ]):
            tk.Label(card, text=label, font=("Segoe UI",9,"bold"),
                     bg=T["PANEL"], fg=T["TEXT"], anchor="w").grid(
                     row=row_i*2, column=0, columnspan=2, sticky="w", pady=(6,1))
            e = tk.Entry(card, textvariable=var, font=("Segoe UI",11),
                         bg=T["ENTRY_BG"], fg=T["ENTRY_FG"],
                         insertbackground=T["TEXT"], relief="flat", bd=4, width=28)
            e.grid(row=row_i*2+1, column=0, columnspan=2, sticky="ew", ipady=5)

        # Loan deduction
        tk.Frame(card, bg=T["SEP"], height=1).grid(
            row=6, column=0, columnspan=2, sticky="ew", pady=(12,6))
        lq = tk.Frame(card, bg=T["PANEL"])
        lq.grid(row=7, column=0, columnspan=2, sticky="ew")
        tk.Label(lq, text="Has loan deduction?", font=("Segoe UI",9),
                 bg=T["PANEL"], fg=T["TEXT"]).pack(side="left")
        tk.Checkbutton(lq, text="Yes", variable=self._has_loan,
                       font=("Segoe UI",9,"bold"), bg=T["PANEL"], fg=T["GOLD"],
                       selectcolor=T["ENTRY_BG"], activebackground=T["PANEL"],
                       activeforeground=T["GOLD"],
                       command=self._toggle_loan).pack(side="right")

        self._loan_frame = tk.Frame(card, bg=T["PANEL"])
        self._loan_frame.grid(row=8, column=0, columnspan=2, sticky="ew")
        tk.Label(self._loan_frame, text="   Monthly loan deduction (TZS):",
                 font=("Segoe UI",8), bg=T["PANEL"], fg=T["GOLD"]).pack(anchor="w")
        self._loan_entry = tk.Entry(self._loan_frame, textvariable=self._loan_var,
                                    font=("Segoe UI",10), bg=T["ENTRY_BG"],
                                    fg=T["ENTRY_FG"], insertbackground=T["TEXT"],
                                    relief="flat", bd=3)
        self._loan_entry.pack(fill="x", ipady=4)
        self._toggle_loan()   # set correct initial state

        # Loan board
        lbq = tk.Frame(card, bg=T["PANEL"])
        lbq.grid(row=9, column=0, columnspan=2, sticky="ew", pady=(8,0))
        tk.Label(lbq, text="Deduct Loan Board? (15% of salary)",
                 font=("Segoe UI",9), bg=T["PANEL"], fg=T["TEXT"]).pack(side="left")
        tk.Checkbutton(lbq, text="Yes", variable=self._has_lb,
                       font=("Segoe UI",9,"bold"), bg=T["PANEL"], fg=T["GOLD"],
                       selectcolor=T["ENTRY_BG"], activebackground=T["PANEL"],
                       activeforeground=T["GOLD"],
                       command=self._preview).pack(side="right")

        # Preview
        tk.Frame(card, bg=T["ACCENT"], height=1).grid(
            row=10, column=0, columnspan=2, sticky="ew", pady=(12,4))
        self._preview_var = tk.StringVar(value="")
        tk.Label(card, textvariable=self._preview_var,
                 font=("Segoe UI",9), bg=T["PANEL"], fg=T["ACCENT"],
                 justify="left").grid(row=11, column=0, columnspan=2, sticky="w")

        # Buttons
        bf = tk.Frame(card, bg=T["PANEL"])
        bf.grid(row=12, column=0, columnspan=2, sticky="ew", pady=(14,0))
        tk.Button(bf, text="💾  Save Changes",
                  font=("Segoe UI",11,"bold"), bg=T["BTN_SAVE"],
                  fg="white", relief="flat", cursor="hand2",
                  command=self._save).pack(side="left", expand=True,
                                           fill="x", padx=(0,4), ipady=9)
        tk.Button(bf, text="✕  Cancel",
                  font=("Segoe UI",11), bg=T["SUBTEXT"],
                  fg="white", relief="flat", cursor="hand2",
                  command=self.destroy).pack(side="left", expand=True,
                                             fill="x", padx=(4,0), ipady=9)

        # Trace changes for live preview
        for var in (self._salary_var, self._allow_var, self._loan_var):
            var.trace_add("write", lambda *_: self._preview())
        self._preview()

    def _toggle_loan(self):
        if self._has_loan.get():
            self._loan_entry.configure(state="normal")
        else:
            self._loan_var.set("")
            self._loan_entry.configure(state="disabled")
        self._preview()

    def _num(self, var):
        try: return float(var.get().replace(",","").strip() or 0)
        except Exception: return 0.0

    def _preview(self, *_):
        s = self._num(self._salary_var)
        a = self._num(self._allow_var)
        if s <= 0:
            self._preview_var.set("Enter a salary to see calculated values.")
            return
        la = self._num(self._loan_var) if self._has_loan.get() else 0.0
        nssf,pay,paye,gross,ld,lb,net,amt = calculate(
            s, a, self._has_loan.get(), la, self._has_lb.get())
        lines = (
            "NSSF: " + "{:,.2f}".format(nssf) +
            "   PAY: " + "{:,.2f}".format(pay) +
            "   PAYE: " + "{:,.2f}".format(paye) + "\n"
            "Net Pay: " + "{:,.2f}".format(net) +
            "   Amount to be Paid: " + "{:,.2f}".format(amt)
        )
        self._preview_var.set(lines)

    def _save(self):
        name   = self._name_var.get().strip()
        salary = self._num(self._salary_var)
        if not name:
            messagebox.showwarning("⚠  Missing Name",
                "Please enter the employee name.", parent=self)
            return
        if salary <= 0:
            messagebox.showwarning("⚠  Missing Salary",
                "Please enter a valid salary.", parent=self)
            return
        la = self._num(self._loan_var) if self._has_loan.get() else 0.0
        if self._has_loan.get() and la <= 0:
            messagebox.showwarning("⚠  Missing Loan Amount",
                "You ticked loan — please enter the deduction amount.", parent=self)
            return
        allowance = self._num(self._allow_var)
        nssf,pay,paye,gross,ld,lb,net,amt = calculate(
            salary, allowance, self._has_loan.get(), la, self._has_lb.get())
        updated = dict(
            name=name, salary=salary, nssf=nssf, pay=pay, paye=paye,
            allowance=allowance, gross_pay=gross, loan_ded=ld,
            loan_board=lb, net_pay=net, amount_paid=amt)
        self.on_save(updated)
        self.destroy()


# ══════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ══════════════════════════════════════════════════════════════
class PayrollApp(tk.Tk):
    def __init__(self, username, role):
        super().__init__()
        self.username         = username
        self.role             = role
        self.records          = []
        self._last_saved_path = None
        self.theme_name       = tk.StringVar(value="Dark Blue")
        self.T                = THEMES["Dark Blue"]
        self.title(APP_TITLE + "  —  " + username)
        try:
            self.state("zoomed")
        except Exception:
            self.attributes("-zoomed", True)
        self.minsize(900, 560)
        self._build_ui()
        self._apply_theme()
        if UPDATE_READY:
            self.after(3000, self._auto_update_check)

    # ── Theme ─────────────────────────────────────────────────
    def _apply_theme(self, *_):
        self.T = THEMES[self.theme_name.get()]
        T = self.T
        self.configure(bg=T["BG"])
        s = ttk.Style(); s.theme_use("clam")
        s.configure("P.Treeview", background=T["TREE_BG"], foreground=T["TREE_FG"],
                    fieldbackground=T["TREE_BG"], rowheight=27, font=("Segoe UI",9))
        s.configure("P.Treeview.Heading", background=T["TREE_HEAD"],
                    foreground="white", font=("Segoe UI",8,"bold"), relief="flat")
        s.map("P.Treeview", background=[("selected", T["ACCENT2"])])
        if hasattr(self, "tree"):
            self.tree.configure(style="P.Treeview")
        self._restyle(self)
        if hasattr(self, "btn_add"):
            self.btn_add.configure(bg=T["BTN_SAVE"])
        if hasattr(self, "btn_remove"):
            self.btn_remove.configure(bg=T["RED"])
        if hasattr(self, "_sf"):
            self._sf.set_bg(T["PANEL"])
        self._refresh_totals_bar()
        # Rebuild the actions menu with updated colours
        if hasattr(self, "_actions_menu"):
            self._build_actions_menu()

    def _restyle(self, root):
        T   = self.T
        ap  = {THEMES[n]["PANEL"]     for n in THEMES}
        ah  = {THEMES[n]["HEADER_BG"] for n in THEMES}
        aa  = {THEMES[n]["ACCENT"]    for n in THEMES}
        as_ = {THEMES[n]["SUBTEXT"]   for n in THEMES}
        ag  = {THEMES[n]["GOLD"]      for n in THEMES}
        ar  = {THEMES[n]["RED"]       for n in THEMES}
        aa2 = {THEMES[n]["ACCENT2"]   for n in THEMES}
        def walk(w):
            cls = type(w).__name__
            try:
                if cls in ("Frame","Canvas"):
                    bg = w.cget("bg")
                    w.configure(bg=T["HEADER_BG"] if bg in ah
                                else T["PANEL"] if bg in ap else T["BG"])
                elif cls == "Label":
                    bg, fg = w.cget("bg"), w.cget("fg")
                    nbg = T["HEADER_BG"] if bg in ah else T["PANEL"] if bg in ap else T["BG"]
                    if   fg in ("white","#ffffff"): nfg = "white"
                    elif fg in aa:   nfg = T["ACCENT"]
                    elif fg in as_:  nfg = T["SUBTEXT"]
                    elif fg in ag:   nfg = T["GOLD"]
                    elif fg in ar:   nfg = T["RED"]
                    elif fg in aa2:  nfg = T["ACCENT2"]
                    else:            nfg = T["TEXT"]
                    w.configure(bg=nbg, fg=nfg)
                elif cls == "Entry":
                    w.configure(bg=T["ENTRY_BG"], fg=T["ENTRY_FG"],
                                insertbackground=T["TEXT"])
                elif cls == "Checkbutton":
                    bg  = w.cget("bg")
                    nbg = T["PANEL"] if bg in ap else T["BG"]
                    w.configure(bg=nbg, fg=T["TEXT"], selectcolor=T["ENTRY_BG"],
                                activebackground=nbg, activeforeground=T["TEXT"])
                elif cls == "OptionMenu":
                    w.configure(bg=T["ENTRY_BG"], fg=T["TEXT"],
                                activebackground=T["ACCENT2"], highlightthickness=0)
                    w["menu"].configure(bg=T["ENTRY_BG"], fg=T["TEXT"],
                                        activebackground=T["ACCENT2"])
            except Exception: pass
            for child in w.winfo_children(): walk(child)
        walk(root)

    # ── Build UI ──────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()

        body = tk.Frame(self, bg=self.T["BG"])
        body.pack(fill="both", expand=True, padx=6, pady=(4,0))
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        # Left: scrollable form, 22% of screen width clamped 220-300 px
        sw = self.winfo_screenwidth()
        panel_w = max(220, min(300, int(sw * 0.22)))
        left_wrap = tk.Frame(body, bg=self.T["PANEL"], width=panel_w)
        left_wrap.grid(row=0, column=0, sticky="nsew", padx=(0,4))
        left_wrap.pack_propagate(False)
        self._sf = ScrollableFrame(left_wrap, bg=self.T["PANEL"])
        self._sf.pack(fill="both", expand=True)

        # Right: table + totals bar
        right_wrap = tk.Frame(body, bg=self.T["PANEL"])
        right_wrap.grid(row=0, column=1, sticky="nsew")
        right_wrap.rowconfigure(0, weight=1)
        right_wrap.columnconfigure(0, weight=1)

        self._build_form(self._sf.inner)
        self._build_table(right_wrap)         # row 0 — expands
        self._build_totals_bar(right_wrap)    # row 1 — fixed height

    # ── Header ────────────────────────────────────────────────
    def _build_header(self):
        T = self.T
        hdr = tk.Frame(self, bg=T["HEADER_BG"])
        hdr.pack(fill="x")

        # ── LEFT: Actions dropdown + Month entry ──────────────
        left = tk.Frame(hdr, bg=T["HEADER_BG"])
        left.pack(side="left", padx=8, pady=6)

        # Big green Actions button
        self._actions_btn = tk.Button(
            left,
            text="☰  Actions  ▾",
            font=("Segoe UI", 10, "bold"),
            bg="#27ae60", fg="white",
            relief="flat", cursor="hand2",
            padx=12, pady=6,
            command=self._show_actions_menu)
        self._actions_btn.pack(side="left")
        ToolTip(self._actions_btn,
                "Save to Excel  |  Print  |  Open Excel  |  New Session")

        # Month entry right next to the button
        tk.Label(left, text="  📅 Month:",
                 font=("Segoe UI",9), bg=T["HEADER_BG"], fg="white").pack(side="left")
        self.month_var = tk.StringVar(
            value=datetime.datetime.now().strftime("%B %Y"))
        tk.Entry(left, textvariable=self.month_var,
                 font=("Segoe UI",9), bg=T["ENTRY_BG"], fg=T["ENTRY_FG"],
                 insertbackground="white", relief="flat", bd=3, width=13
                 ).pack(side="left", padx=(4,0), ipady=3)

        # App title (centre)
        mid = tk.Frame(hdr, bg=T["HEADER_BG"])
        mid.pack(side="left", expand=True)
        tk.Label(mid, text="💼  PAYROLL PROCESSOR",
                 font=("Segoe UI",14,"bold"),
                 bg=T["HEADER_BG"], fg="white").pack()
        tk.Label(mid,
                 text="Tanzania  ·  TRA PAYE 2024/25  ·  v" + VERSION + "  ·  " + self.username + " (" + self.role + ")",
                 font=("Segoe UI",8), bg=T["HEADER_BG"], fg="white").pack()

        # ── RIGHT: Theme, Users, Password, Update, Logout ─────
        right = tk.Frame(hdr, bg=T["HEADER_BG"])
        right.pack(side="right", padx=8)

        tk.Label(right, text="🎨", font=("Segoe UI",9),
                 bg=T["HEADER_BG"], fg="white").pack(side="left")
        om = tk.OptionMenu(right, self.theme_name, *THEMES.keys(),
                           command=self._apply_theme)
        om.configure(bg=T["ENTRY_BG"], fg=T["TEXT"], font=("Segoe UI",8),
                     relief="flat", highlightthickness=0, padx=4)
        om["menu"].configure(bg=T["ENTRY_BG"], fg=T["TEXT"], font=("Segoe UI",8))
        om.pack(side="left", padx=(2,8))

        if self.role == "admin":
            b = tk.Button(right, text="👥 Users",
                          font=("Segoe UI",8,"bold"), bg=T["BTN_SAVE"],
                          fg="white", relief="flat", cursor="hand2",
                          command=lambda: UserManagerDialog(self, self.T))
            b.pack(side="left", padx=2, ipady=4, ipadx=6)
            ToolTip(b, "Add or remove users")

        b2 = tk.Button(right, text="🔒 Password",
                       font=("Segoe UI",8,"bold"), bg=T["ENTRY_BG"],
                       fg=T["TEXT"], relief="flat", cursor="hand2",
                       command=lambda: ChangePasswordDialog(self, self.T, self.username))
        b2.pack(side="left", padx=2, ipady=4, ipadx=6)
        ToolTip(b2, "Change your password")

        if UPDATE_READY:
            b3 = tk.Button(right, text="⬆ Update",
                           font=("Segoe UI",8,"bold"), bg="#16a085",
                           fg="white", relief="flat", cursor="hand2",
                           command=self._manual_update_check)
            b3.pack(side="left", padx=2, ipady=4, ipadx=6)
            ToolTip(b3, "Check for a newer version")

        b4 = tk.Button(right, text="⏻ Logout",
                       font=("Segoe UI",8,"bold"), bg=T["RED"],
                       fg="white", relief="flat", cursor="hand2",
                       command=self._logout)
        b4.pack(side="left", padx=(2,0), ipady=4, ipadx=6)
        ToolTip(b4, "Log out")

        # Build the dropdown menu object
        self._build_actions_menu()

    def _build_actions_menu(self):
        """Build (or rebuild) the tk.Menu used by the Actions button."""
        T = self.T
        m = tk.Menu(self, tearoff=0,
                    bg=T["MENU_BG"], fg=T["MENU_FG"],
                    activebackground=T["MENU_ACT"],
                    activeforeground="white",
                    font=("Segoe UI", 10),
                    relief="flat", bd=1)

        m.add_command(
            label="  💾   Save  (same file)",
            command=self.save_excel)
        m.add_command(
            label="  💾   Save As  (choose new location)",
            command=self.save_excel_as)
        m.add_command(
            label="  🖨    Print Payroll",
            command=self.print_payroll)
        m.add_command(
            label="  📂   Open Last Excel",
            command=self.open_file)
        m.add_separator()
        m.add_command(
            label="  🔄   New Session  (clear list)",
            command=self.new_session)

        self._actions_menu = m

    def _show_actions_menu(self):
        """Pop the menu directly below the Actions button."""
        btn = self._actions_btn
        x   = btn.winfo_rootx()
        y   = btn.winfo_rooty() + btn.winfo_height()
        self._actions_menu.tk_popup(x, y)

    # ── Form (left panel) ─────────────────────────────────────
    def _build_form(self, p):
        T = self.T
        tk.Label(p, text="➕  ADD EMPLOYEE",
                 font=("Segoe UI",10,"bold"), bg=T["PANEL"], fg=T["ACCENT"]
                 ).pack(pady=(10,2), padx=10, anchor="w")
        tk.Frame(p, bg=T["ACCENT"], height=2).pack(fill="x", padx=10)

        self.entries = {}
        self._big_field(p, "👤  Employee Name",      "Name",      "e.g. John Banda")
        self._big_field(p, "💰  Basic Salary (TZS)",  "Salary",    "e.g. 500000")
        self._big_field(p, "➕  Allowance (TZS)",     "ALLOWANCE", "Leave as 0 if none")

        note = tk.Frame(p, bg=T["ENTRY_BG"])
        note.pack(fill="x", padx=10, pady=(4,0))
        tk.Label(note, text="ℹ  P.A.Y.E auto-calculated (TRA 2024/25)",
                 font=("Segoe UI",7,"italic"), bg=T["ENTRY_BG"],
                 fg=T["SUBTEXT"], padx=6, pady=4).pack(anchor="w")

        # Loan deduction
        tk.Frame(p, bg=T["SEP"], height=1).pack(fill="x", padx=10, pady=8)
        self.has_loan = tk.BooleanVar()
        lq = tk.Frame(p, bg=T["PANEL"]); lq.pack(fill="x", padx=10, pady=2)
        tk.Label(lq, text="Employee has a loan?",
                 font=("Segoe UI",9), bg=T["PANEL"], fg=T["TEXT"]).pack(side="left")
        tk.Checkbutton(lq, text="Yes", variable=self.has_loan,
                       font=("Segoe UI",9,"bold"), bg=T["PANEL"], fg=T["GOLD"],
                       selectcolor=T["ENTRY_BG"], activebackground=T["PANEL"],
                       activeforeground=T["GOLD"],
                       command=self._toggle_loan).pack(side="right")
        self.loan_amount_frame = tk.Frame(p, bg=T["PANEL"])
        self.loan_amount_frame.pack(fill="x", padx=10, pady=(0,4))
        tk.Label(self.loan_amount_frame, text="   Monthly loan deduction (TZS):",
                 font=("Segoe UI",8), bg=T["PANEL"], fg=T["GOLD"]).pack(anchor="w")
        self.loan_var   = tk.StringVar()
        self.loan_entry = tk.Entry(self.loan_amount_frame, textvariable=self.loan_var,
                                   font=("Segoe UI",11), bg=T["ENTRY_BG"],
                                   fg=T["ENTRY_FG"], insertbackground=T["TEXT"],
                                   relief="flat", bd=4, state="disabled")
        self.loan_entry.pack(fill="x", ipady=5, pady=(2,0))

        # Loan board
        tk.Frame(p, bg=T["SEP"], height=1).pack(fill="x", padx=10, pady=6)
        self.has_loan_board = tk.BooleanVar()
        lbq = tk.Frame(p, bg=T["PANEL"]); lbq.pack(fill="x", padx=10, pady=2)
        tk.Label(lbq, text="Deduct Loan Board? (15% of salary)",
                 font=("Segoe UI",9), bg=T["PANEL"], fg=T["TEXT"]).pack(side="left")
        tk.Checkbutton(lbq, text="Yes", variable=self.has_loan_board,
                       font=("Segoe UI",9,"bold"), bg=T["PANEL"], fg=T["GOLD"],
                       selectcolor=T["ENTRY_BG"], activebackground=T["PANEL"],
                       activeforeground=T["GOLD"],
                       command=self.preview_calc).pack(side="right")
        self.lb_info = tk.Label(p, text="   15% will be deducted from salary",
                                 font=("Segoe UI",7,"italic"),
                                 bg=T["PANEL"], fg=T["SUBTEXT"])
        self.lb_info.pack(anchor="w", padx=14)

        # Calculated preview
        tk.Frame(p, bg=T["ACCENT"], height=1).pack(fill="x", padx=10, pady=(10,3))
        tk.Label(p, text="📊  CALCULATED VALUES",
                 font=("Segoe UI",8,"bold"), bg=T["PANEL"], fg=T["SUBTEXT"]
                 ).pack(anchor="w", padx=10)
        self.calc_vars = {}
        for af, ck, tip in [
            ("NSSF 10%",          "GOLD",    "10% of basic salary"),
            ("PAY",               "GOLD",    "Salary minus NSSF"),
            ("P.A.Y.E",           "RED",     "Income tax (TRA brackets)"),
            ("GROSS PAY",         "GOLD",    "Salary + Allowance"),
            ("LOAN BOARD",        "SUBTEXT", "15% if ticked above"),
            ("NET PAY",           "ACCENT",  "After NSSF and PAYE"),
            ("AMOUNT TO BE PAID", "ACCENT",  "What employee receives in hand"),
        ]:
            f = tk.Frame(p, bg=T["PANEL"]); f.pack(fill="x", padx=10, pady=1)
            lbl = tk.Label(f, text=af, width=20, anchor="w",
                           font=("Segoe UI",8), bg=T["PANEL"], fg=T["SUBTEXT"])
            lbl.pack(side="left")
            ToolTip(lbl, tip)
            var = tk.StringVar(value="—"); self.calc_vars[af] = var
            tk.Label(f, textvariable=var, anchor="e",
                     font=("Segoe UI",9,"bold"), bg=T["PANEL"], fg=T[ck]).pack(side="right")

        for key in ("Salary", "ALLOWANCE"):
            self.entries[key].trace_add("write", lambda *_: self.preview_calc())
        self.loan_var.trace_add("write", lambda *_: self.preview_calc())

        tk.Frame(p, bg=T["ACCENT"], height=2).pack(fill="x", padx=10, pady=(10,4))
        self.btn_add = tk.Button(p, text="➕   ADD THIS EMPLOYEE",
                                  font=("Segoe UI",11,"bold"),
                                  bg=T["BTN_SAVE"], fg="white", relief="flat",
                                  cursor="hand2", command=self.add_employee)
        self.btn_add.pack(fill="x", padx=10, ipady=10, pady=3)
        ToolTip(self.btn_add, "Click to add this employee to the payroll list")

        tk.Button(p, text="🗑  Clear form",
                  font=("Segoe UI",8), bg=T["PANEL"], fg=T["SUBTEXT"],
                  relief="flat", cursor="hand2",
                  command=self.clear_form).pack(pady=(0,12))

    def _big_field(self, parent, label, key, hint=""):
        T = self.T
        f = tk.Frame(parent, bg=T["PANEL"]); f.pack(fill="x", padx=10, pady=4)
        tk.Label(f, text=label, anchor="w", font=("Segoe UI",9,"bold"),
                 bg=T["PANEL"], fg=T["TEXT"]).pack(anchor="w")
        var = tk.StringVar()
        tk.Entry(f, textvariable=var, font=("Segoe UI",11),
                 bg=T["ENTRY_BG"], fg=T["ENTRY_FG"],
                 insertbackground=T["TEXT"], relief="flat", bd=4).pack(fill="x", ipady=6)
        if hint:
            tk.Label(f, text="   " + hint, font=("Segoe UI",7),
                     bg=T["PANEL"], fg=T["SUBTEXT"]).pack(anchor="w")
        self.entries[key] = var

    # ── Table (right panel, row 0) ────────────────────────────
    def _build_table(self, p):
        T = self.T
        p.rowconfigure(0, weight=1)
        p.columnconfigure(0, weight=1)

        hrow = tk.Frame(p, bg=T["PANEL"])
        hrow.grid(row=0, column=0, sticky="new")
        tk.Label(hrow, text="📋  PAYROLL REGISTER",
                 font=("Segoe UI",10,"bold"), bg=T["PANEL"], fg=T["ACCENT"]
                 ).pack(side="left", padx=10, pady=(8,2))
        self.count_var = tk.StringVar(value="No employees yet")
        tk.Label(hrow, textvariable=self.count_var,
                 font=("Segoe UI",8), bg=T["PANEL"], fg=T["SUBTEXT"]
                 ).pack(side="right", padx=10, pady=(8,2))
        tk.Label(hrow, text="✏  Double-click any row to edit",
                 font=("Segoe UI",7,"italic"), bg=T["PANEL"], fg=T["SUBTEXT"]
                 ).pack(side="right", padx=10, pady=(8,2))
        tk.Frame(p, bg=T["ACCENT"], height=2).grid(
            row=0, column=0, sticky="sew", padx=8)

        tf = tk.Frame(p, bg=T["PANEL"])
        tf.grid(row=0, column=0, sticky="nsew", padx=4, pady=(32,0))
        tf.columnconfigure(0, weight=1)
        tf.rowconfigure(0, weight=1)

        s = ttk.Style(); s.theme_use("clam")
        s.configure("P.Treeview", background=T["TREE_BG"], foreground=T["TREE_FG"],
                    fieldbackground=T["TREE_BG"], rowheight=27, font=("Segoe UI",9))
        s.configure("P.Treeview.Heading", background=T["TREE_HEAD"],
                    foreground="white", font=("Segoe UI",8,"bold"), relief="flat")
        s.map("P.Treeview", background=[("selected", T["ACCENT2"])])

        self.tree = ttk.Treeview(tf, columns=HEADERS, show="headings",
                                  style="P.Treeview", height=15)

        # Column widths — proportional weights (Name gets 2x, others 1x)
        # Actual pixel widths are set dynamically on window resize
        self._col_weights = {
            "Name": 2.2, "Salary": 1.1, "NSSF 10%": 1.0, "PAY": 1.0,
            "P.A.Y.E": 1.0, "ALLOWANCE": 1.1, "GROSS PAY": 1.1,
            "LOAN Deduction": 1.2, "LOAN BOARD (15%)": 1.2,
            "NET PAY": 1.1, "AMOUNT TO BE PAID": 1.4,
        }
        for c in HEADERS:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=80, anchor="center",
                             minwidth=50, stretch=True)
        self.tree.column("Name", anchor="w")

        # Vertical scroll only — columns will always fit horizontally
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        # Recalculate column widths whenever the table frame is resized
        tf.bind("<Configure>", self._fit_columns)
        # Double-click a row to edit
        self.tree.bind("<Double-1>", self._on_row_double_click)

        rm = tk.Frame(p, bg=T["PANEL"])
        rm.grid(row=0, column=0, sticky="se", padx=6, pady=2)
        self.btn_remove = tk.Button(rm, text="🗑  Remove Selected Row",
                                     font=("Segoe UI",8),
                                     bg=T["RED"], fg="white", relief="flat",
                                     cursor="hand2", command=self.remove_selected)
        self.btn_remove.pack(ipady=3, ipadx=8)
        ToolTip(self.btn_remove, "Click a row first, then click here to remove it")

    def _fit_columns(self, event=None):
        """Distribute available width among columns proportionally."""
        if not hasattr(self, "tree"):
            return
        try:
            total_w = self.tree.winfo_width() - 20   # subtract scrollbar width
            if total_w <= 0:
                return
            total_weight = sum(self._col_weights.values())
            for col in HEADERS:
                w = int(total_w * self._col_weights[col] / total_weight)
                self.tree.column(col, width=max(w, 50))
        except Exception:
            pass

    # ── Totals bar (right panel, row 1) ───────────────────────
    def _build_totals_bar(self, parent):
        T = self.T
        self.totals_bar = tk.Frame(parent, bg=T["PANEL"], pady=5)
        self.totals_bar.grid(row=1, column=0, sticky="ew", padx=4, pady=(0,4))

        tk.Label(self.totals_bar, text="📊 TOTALS:",
                 font=("Segoe UI",8,"bold"), bg=T["PANEL"], fg=T["SUBTEXT"]
                 ).pack(side="left", padx=(8,4))

        self._tot_vars = {}
        for label, key, ck in [
            ("Staff",             "count",        "SUBTEXT"),
            ("Total Salary",      "total_salary", "GOLD"),
            ("NSSF",              "total_nssf",   "SUBTEXT"),
            ("PAYE Tax",          "total_paye",   "RED"),
            ("Net Pay",           "total_net",    "GOLD"),
            ("AMOUNT TO PAY OUT", "total_amount", "ACCENT"),
        ]:
            box = tk.Frame(self.totals_bar, bg=T["ENTRY_BG"], padx=8, pady=3)
            box.pack(side="left", padx=3)
            tk.Label(box, text=label, font=("Segoe UI",6),
                     bg=T["ENTRY_BG"], fg=T["SUBTEXT"]).pack()
            var = tk.StringVar(value="—"); self._tot_vars[key] = var
            tk.Label(box, textvariable=var, font=("Segoe UI",9,"bold"),
                     bg=T["ENTRY_BG"], fg=T[ck]).pack()

    def _refresh_totals_bar(self):
        if not hasattr(self, "_tot_vars"): return
        if not self.records:
            for v in self._tot_vars.values(): v.set("—")
            return
        self._tot_vars["count"].set(str(len(self.records)))
        self._tot_vars["total_salary"].set("{:,.0f}".format(sum(r["salary"]      for r in self.records)))
        self._tot_vars["total_nssf"].set("{:,.0f}".format(sum(r["nssf"]          for r in self.records)))
        self._tot_vars["total_paye"].set("{:,.0f}".format(sum(r["paye"]          for r in self.records)))
        self._tot_vars["total_net"].set("{:,.0f}".format(sum(r["net_pay"]        for r in self.records)))
        self._tot_vars["total_amount"].set("{:,.0f}".format(sum(r["amount_paid"] for r in self.records)))
        T = self.T
        for child in self.totals_bar.winfo_children():
            if isinstance(child, tk.Frame):
                child.configure(bg=T["ENTRY_BG"])
                for lbl in child.winfo_children():
                    if isinstance(lbl, tk.Label): lbl.configure(bg=T["ENTRY_BG"])

    # ── Logic ─────────────────────────────────────────────────
    def _toggle_loan(self):
        if self.has_loan.get():
            self.loan_entry.configure(state="normal")
            self.loan_entry.focus_set()
        else:
            self.loan_var.set("")
            self.loan_entry.configure(state="disabled")
        self.preview_calc()

    def _num(self, key):
        try: return float(self.entries[key].get().replace(",","").strip() or 0)
        except Exception: return 0.0

    def _loan_amt(self):
        try: return float(self.loan_var.get().replace(",","") or 0)
        except Exception: return 0.0

    def preview_calc(self, *_):
        s = self._num("Salary"); a = self._num("ALLOWANCE")
        nssf,pay,paye,gross,ld,lb,net,amt = calculate(
            s, a, self.has_loan.get(), self._loan_amt(), self.has_loan_board.get())
        fmt = lambda v: "TZS  " + "{:,.2f}".format(v) if s else "—"
        self.calc_vars["NSSF 10%"].set(fmt(nssf))
        self.calc_vars["PAY"].set(fmt(pay))
        self.calc_vars["P.A.Y.E"].set(fmt(paye))
        self.calc_vars["GROSS PAY"].set(fmt(gross))
        self.calc_vars["LOAN BOARD"].set(fmt(lb) if self.has_loan_board.get() else "Not selected")
        self.calc_vars["NET PAY"].set(fmt(net))
        self.calc_vars["AMOUNT TO BE PAID"].set(fmt(amt))
        self.lb_info.configure(
            text="   Deduction = TZS " + "{:,.2f}".format(lb)
            if self.has_loan_board.get() and s
            else "   15% will be deducted from salary")

    def add_employee(self):
        name   = self.entries["Name"].get().strip()
        salary = self._num("Salary")
        if not name:
            messagebox.showwarning("⚠  Missing Name",
                "Please type the employee's full name.")
            return
        if salary <= 0:
            messagebox.showwarning("⚠  Missing Salary",
                "Please enter the employee's salary.")
            return
        if self.has_loan.get() and self._loan_amt() <= 0:
            messagebox.showwarning("⚠  Missing Loan Amount",
                "You ticked 'Has loan?' — please enter the monthly deduction amount.")
            return
        allowance = self._num("ALLOWANCE")
        nssf,pay,paye,gross,ld,lb,net,amt = calculate(
            salary, allowance, self.has_loan.get(),
            self._loan_amt(), self.has_loan_board.get())
        rec = dict(name=name, salary=salary, nssf=nssf, pay=pay, paye=paye,
                   allowance=allowance, gross_pay=gross, loan_ded=ld,
                   loan_board=lb, net_pay=net, amount_paid=amt)
        self.records.append(rec)
        self.tree.insert("", "end", values=(
            name,
            "{:,.2f}".format(salary),  "{:,.2f}".format(nssf),
            "{:,.2f}".format(pay),     "{:,.2f}".format(paye),
            "{:,.2f}".format(allowance), "{:,.2f}".format(gross),
            "{:,.2f}".format(ld),      "{:,.2f}".format(lb),
            "{:,.2f}".format(net),     "{:,.2f}".format(amt),
        ))
        n = len(self.records)
        self.count_var.set(str(n) + " employee" + ("s" if n != 1 else "") + " added")
        self._refresh_totals_bar()
        # Refit columns after new row added
        self.after(50, self._fit_columns)
        self.clear_form()
        messagebox.showinfo("✅  Added",
            name + " has been added.\n\n"
            "Amount to be paid:  TZS " + "{:,.2f}".format(amt) + "\n\n"
            "Total employees: " + str(n))

    def _on_row_double_click(self, event):
        """Open EditEmployeeDialog for the clicked row."""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        idx = self.tree.index(item)
        rec = self.records[idx]

        def on_save(updated):
            # Update internal record
            self.records[idx] = updated
            # Update tree row
            self.tree.item(item, values=(
                updated["name"],
                "{:,.2f}".format(updated["salary"]),
                "{:,.2f}".format(updated["nssf"]),
                "{:,.2f}".format(updated["pay"]),
                "{:,.2f}".format(updated["paye"]),
                "{:,.2f}".format(updated["allowance"]),
                "{:,.2f}".format(updated["gross_pay"]),
                "{:,.2f}".format(updated["loan_ded"]),
                "{:,.2f}".format(updated["loan_board"]),
                "{:,.2f}".format(updated["net_pay"]),
                "{:,.2f}".format(updated["amount_paid"]),
            ))
            self._refresh_totals_bar()
            messagebox.showinfo("✅  Updated",
                updated["name"] + " has been updated successfully.")

        EditEmployeeDialog(self, self.T, rec, on_save)

    def remove_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Nothing selected",
                "Please click on an employee row first.")
            return
        if messagebox.askyesno("Confirm", "Remove the selected employee?"):
            for item in sel:
                idx = self.tree.index(item)
                self.tree.delete(item)
                self.records.pop(idx)
            n = len(self.records)
            self.count_var.set(
                str(n) + " employee" + ("s" if n != 1 else "") + " added"
                if n else "No employees yet")
            self._refresh_totals_bar()

    def clear_form(self):
        for v in self.entries.values(): v.set("")
        self.has_loan.set(False)
        self.loan_var.set("")
        self.loan_entry.configure(state="disabled")
        self.has_loan_board.set(False)
        for v in self.calc_vars.values(): v.set("—")
        self.lb_info.configure(text="   15% will be deducted from salary")

    def save_excel(self):
        if not self.records:
            messagebox.showwarning("⚠  Nothing to Save",
                "Please add at least one employee first.")
            return
        # If we already have a path for this session, save directly — no dialog
        if self._last_saved_path:
            fp = self._last_saved_path
        else:
            month = self.month_var.get().replace(" ", "_")
            fp = filedialog.asksaveasfilename(
                title="Choose where to save the Excel file",
                defaultextension=".xlsx",
                filetypes=[("Excel file", "*.xlsx")],
                initialfile="Payroll_" + month + ".xlsx")
            if not fp:
                return
        try:
            save_to_excel(self.records, fp, self.username, self.month_var.get())
            self._last_saved_path = fp
            messagebox.showinfo("✅  Saved",
                "Payroll saved to:\n\n" + fp + "\n\n"
                + str(len(self.records)) + " employee(s) included.")
        except Exception as e:
            messagebox.showerror("❌  Error saving", str(e))

    def save_excel_as(self):
        """Force a Save As dialog — lets user pick a new location."""
        if not self.records:
            messagebox.showwarning("⚠  Nothing to Save",
                "Please add at least one employee first.")
            return
        month = self.month_var.get().replace(" ", "_")
        fp = filedialog.asksaveasfilename(
            title="Save As — choose location",
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")],
            initialfile="Payroll_" + month + ".xlsx")
        if not fp:
            return
        try:
            save_to_excel(self.records, fp, self.username, self.month_var.get())
            self._last_saved_path = fp
            messagebox.showinfo("✅  Saved",
                "Payroll saved to:\n\n" + fp)
        except Exception as e:
            messagebox.showerror("❌  Error saving", str(e))

    def print_payroll(self):
        if not self.records:
            messagebox.showwarning("⚠  Nothing to Print",
                "Please add at least one employee first.")
            return
        html = make_print_html(self.records, self.month_var.get(), self.username)
        open_print_in_browser(html)
        messagebox.showinfo("🖨  How to Print",
            "The payroll has opened in your web browser.\n\n"
            "To print:\n\n"
            "  1.  Look at the browser window that opened\n"
            "  2.  Hold  Ctrl  and press  P\n"
            "  3.  Choose your printer and click  Print\n\n"
            "Tip: choose  'Save as PDF'  to get a PDF file.")

    def open_file(self):
        if not self._last_saved_path:
            messagebox.showinfo("No file yet",
                "You have not saved an Excel file yet.\n\n"
                "Use  ☰ Actions → Save to Excel  first.")
            return
        if not os.path.exists(self._last_saved_path):
            messagebox.showinfo("File not found",
                "Could not find:\n" + self._last_saved_path)
            return
        try: os.startfile(self._last_saved_path)
        except AttributeError: subprocess.call(["xdg-open", self._last_saved_path])

    def new_session(self):
        if self.records and not messagebox.askyesno(
                "🔄  New Session",
                "This will clear all employees from the list.\n\n"
                "⚠  Make sure you have already saved the Excel file!\n\nContinue?"):
            return
        self.records.clear()
        for item in self.tree.get_children(): self.tree.delete(item)
        self.count_var.set("No employees yet")
        self._refresh_totals_bar()
        self.clear_form()
        self._last_saved_path = None

    # ── Update methods ────────────────────────────────────────
    def _run_update_check(self, on_complete):
        """
        Fetch GitHub code in background.
        on_complete called on main thread with "same" | "error" | <new_source>.
        """
        holder = [None]
        fetch_update_in_bg(holder)
        def poll(elapsed=0):
            if elapsed >= 12000:
                on_complete("error")
                return
            val = holder[0]
            if val is None:
                self.after(400, lambda: poll(elapsed + 400))
            else:
                on_complete(val)
        self.after(400, lambda: poll(400))

    def _auto_update_check(self):
        """Called silently on every login. Prompts only if change detected."""
        def on_done(result):
            if result not in ("same", "error"):
                # Change detected — ask user
                self.after(0, lambda: self._prompt_update(result))
        self._run_update_check(on_done)

    def _manual_update_check(self):
        """User clicked the Update button — show result either way."""
        # Disable the button while checking
        for child in self.winfo_children():
            pass   # can't easily find btn here, just let it run

        def on_done(result):
            if result == "same":
                messagebox.showinfo("✅  Up to Date",
                    "Your app is identical to the version on GitHub.\n"
                    "No update needed.")
            elif result == "error":
                messagebox.showwarning("⚠  Could Not Check",
                    "Could not reach GitHub.\n"
                    "Check your internet connection and try again.")
            else:
                self._prompt_update(result)
        self._run_update_check(on_done)

    def _prompt_update(self, new_source):
        """Ask the user whether to update now or later."""
        # Custom dialog with three choices: Update Now / Later / Skip
        dlg = tk.Toplevel(self)
        dlg.title("🔄  Update Available")
        dlg.configure(bg="#1e2a3a")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        dlg.geometry("420x240+" + str((sw-420)//2) + "+" + str((sh-240)//2))

        tk.Label(dlg, text="🔄  Update Available",
                 font=("Segoe UI",14,"bold"), bg="#1e2a3a", fg="#2ecc71").pack(pady=(22,4))
        tk.Label(dlg,
                 text="A newer version of this app has been found on GitHub.\n"
                      "Would you like to update now?",
                 font=("Segoe UI",9), bg="#1e2a3a", fg="#ecf0f1",
                 justify="center").pack(pady=(0,18))

        def do_update():
            dlg.destroy()
            apply_update(new_source, self)

        def do_later():
            dlg.destroy()
            messagebox.showinfo("Update Later",
                "No problem! You can update later using the  ⬆ Update  button.")

        bf = tk.Frame(dlg, bg="#1e2a3a")
        bf.pack()
        tk.Button(bf, text="✅  Update Now  (app will restart)",
                  font=("Segoe UI",10,"bold"), bg="#27ae60", fg="white",
                  relief="flat", cursor="hand2", padx=10, pady=8,
                  command=do_update).pack(side="left", padx=8)
        tk.Button(bf, text="🕐  Later",
                  font=("Segoe UI",10), bg="#2d4059", fg="#ecf0f1",
                  relief="flat", cursor="hand2", padx=10, pady=8,
                  command=do_later).pack(side="left", padx=8)

    def _logout(self):
        if messagebox.askyesno("⏻  Logout",
                "Log out?\n\n⚠  Save your work first!"):
            self.destroy()
            start_app()

# ══════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════
def start_app():
    # First ever launch — no user accounts exist yet
    if not has_any_users():
        setup = FirstSetupWindow()
        setup.mainloop()
        if not setup.created_user:
            return  # user closed the window without creating account
        # Log them in directly — no need to type it again
        app = PayrollApp(setup.created_user, setup.created_role)
        app.mainloop()
        return

    # Normal launch — show login screen
    login = LoginWindow()
    login.mainloop()
    if login.logged_in_user:
        app = PayrollApp(login.logged_in_user, login.logged_in_role)
        app.mainloop()

if __name__ == "__main__":
    start_app()